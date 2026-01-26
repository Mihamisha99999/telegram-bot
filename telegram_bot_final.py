#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Telegram Bot для управления командой создателей контента
Отслеживает создание видео, загрузки, расчет доходов и выплаты зарплаты
ВЕРСИЯ С 6 ДЕВУШКАМИ (добавлены ☀️ и 🌸)
"""

import os
import json
import random
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ConversationHandler, filters, ContextTypes

# Установка необходимых библиотек:
# pip install python-telegram-bot openpyxl

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Библиотека openpyxl не установлена. Установите: pip install openpyxl")

# Состояния для ConversationHandler (УНИКАЛЬНЫЕ для каждого handler)
VIDEO_TYPE, VIDEO_NAME = range(2)
UPLOAD_COUNT = 100
PAY_USER, PAY_AMOUNT = range(200, 202)
BROADCAST_MSG = 300

# ✅ КОНФИГУРАЦИЯ ПОЛЬЗОВАТЕЛЕЙ - 6 ДЕВУШЕК (4 старые + 2 новые)
USERS_CONFIG = {
    "Вика": {
        "role": "creator_uploader",
        "rates": {"a2e": 350, "makefilm": 450, "grok": 550, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Юля": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 500, "grok": 550, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Алла": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Аня": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "☀️": {  # 🔥 НОВАЯ ДЕВУШКА - СОЛНЫШКО
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "🌸": {  # 🔥 НОВАЯ ДЕВУШКА - ЦВЕТОЧЕК
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    }
}

# Админы (вы и муж)
ADMINS = {
    "admin": 2147091471,
    "husband": 2106439695
}

# Мотивационные фразы
MOTIVATIONAL_PHRASES = {
    "video_created": [
        "{name}, ты королева! 👑 Ещё одно видео {video_type}! +{amount} грн 💰 Так держать! 🔥",
        "{name}, ты звезда! ⭐ Видео {video_type} создано! +{amount} грн 💸 Ты лучшая! 💪",
        "Вау, {name}! Ещё одно видео! 🚀 +{amount} грн 💎 Ты просто огонь! 🔥",
        "Супер, {name}! Ты на волне! 🌊 Видео {video_type} готово! +{amount} грн 💰",
        "{name}, ты чемпион! 🏆 Ещё одно видео {video_type}! +{amount} грн 🎉",
        "Красавица, {name}! 💃 Видео {video_type} создано! +{amount} грн ✨",
        "{name}, ты машина! 🚗💨 Ещё одно видео! +{amount} грн 💪",
        "Браво, {name}! 👏 Видео {video_type} готово! +{amount} грн 🌟"
    ],
    "video_uploaded": [
        "{name}, ты ж моя звездачка! ⭐ {count} видео загружено! +{amount} грн 💸 Ты лучшая! 💪",
        "Супер, {name}! 🎉 {count} видео на сайтах! +{amount} грн 💰 Так держать! 🔥",
        "{name}, ты королева загрузок! 👑 {count} видео выставлено! +{amount} грн 💎",
        "Вау, {name}! 🚀 {count} видео загружено! +{amount} грн ✨ Ты просто огонь! 🔥",
        "{name}, ты чемпион! 🏆 {count} видео на сайтах! +{amount} грн 🎊",
        "Красотка, {name}! 💃 {count} видео выставлено! +{amount} грн 🌟"
    ],
    "payment": [
        "🎉 {name}, тебе выплачено {amount} грн! 💰 Спасибо за отличную работу! 🌟",
        "🎊 {name}, зарплата готова! 💸 {amount} грн! Ты лучшая! 👑",
        "💰 {name}, выплата произведена! {amount} грн! Спасибо за труд! 🙏",
        "🌟 {name}, получи свои {amount} грн! 💵 Ты супер! 🔥",
        "🎁 {name}, твои {amount} грн ждут тебя! 💰 Продолжай в том же духе! 💪"
    ],
    "husband_jokes": [
        "Отлично. Теперь переделай.💪",
        "Это было хорошо. Почти",
        "Спасибо, теперь я знаю, чего избегать.",
        "Это решение или крик о помощи? 😄",
        "Интересно. Бесполезно, но интересно.💎 Браво! 👏",
        "Команда мечты! 🌟 Вы все супергерои! 🦸‍♀️",
        "Сегодня рекорды будут побиты! 🚀 Вперёд! 💪",
        "Это не провал. Это стиль. Плохой.🔥",
        "Хорошая попытка.",
        "😬 Это выглядело лучше в твоей голове, да? 😬",
        "🫣 Ты опять превзошёл ожидания. Самые низкие. 🫣",
        "🤷‍♀️ Я вижу усилия. Результат где? 🤷‍♀️",
        "😌 Ну что ж. Опыт получен. Живём дальше. 😌",
        "💥 Это был план или импровизированная катастрофа? 💥",
        "🙃 У тебя талант. Жаль, не в эту сторону. 🙃",
        "🔥 Ничего, опыт — лучший учитель. Особенно через боль. 🔥",
        "🧩 Ты почти понял. Осталось всё остальное. 🧩",
        "🎭 Давай сделаем вид, что это было задумано. 🎭",
        "👀 Я вижу старание. Результат — нет. 👀",
        "🧪 Это не провал. Это демо-версия провала. 🧪",
        "🔍 Ты вдохновляешь меня… перепроверять всё. 🔍",
        "🎯 Хорошая попытка. Следующая тоже будет не она. 🎯"
    ]
}

DATABASE_FILE = "bot_database.json"

def load_database():
    if os.path.exists(DATABASE_FILE):
        with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "users": USERS_CONFIG,
        "admins": ADMINS,
        "videos": [],
        "uploads": [],
        "payments": [],
        "balances": {name: 0 for name in USERS_CONFIG.keys()}
    }

def save_database(db):
    with open(DATABASE_FILE, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

db = load_database()

def get_user_by_telegram_id(telegram_id):
    for name, config in db["users"].items():
        if config.get("telegram_id") == telegram_id:
            return name
    return None

def is_admin(telegram_id):
    return telegram_id in [db["admins"]["admin"], db["admins"]["husband"]]

def get_user_keyboard(user_name):
    keyboard = []
    keyboard.append([KeyboardButton("🎬 Создала видео")])
    keyboard.append([KeyboardButton("📤 Загрузила видео")])
    keyboard.append([KeyboardButton("💰 Мой доход"), KeyboardButton("📊 Моя статистика")])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_admin_keyboard():
    keyboard = [
        [KeyboardButton("📊 Полная статистика"), KeyboardButton("⚙️ Текущий баланс")],
        [KeyboardButton("💸 Выплатить зарплату"), KeyboardButton("📈 История выплат")],
        [KeyboardButton("🎬 Все видео"), KeyboardButton("📊 Экспорт в Excel")],
        [KeyboardButton("📢 Срочное сообщение всем")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_husband_keyboard():
    keyboard = [
        [KeyboardButton("📊 Полная статистика"), KeyboardButton("⚙️ Текущий баланс")],
        [KeyboardButton("😂 Отправить прикольчик"), KeyboardButton("💪 Мотивация девочкам")],
        [KeyboardButton("🎬 Все видео"), KeyboardButton("📈 История выплат")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if user_name:
        keyboard = get_user_keyboard(user_name)
        await update.message.reply_text(
            f"Привет, {user_name}! 👋\n\n"
            f"Я помогу тебе отслеживать твою работу и доход! 💰\n\n"
            f"Используй кнопки ниже для работы с ботом:",
            reply_markup=keyboard
        )
    elif is_admin(telegram_id):
        if telegram_id == db["admins"]["husband"]:
            keyboard = get_husband_keyboard()
            await update.message.reply_text(
                f"Привет, босс! 👨‍💼\n\n"
                f"Ты можешь видеть всю статистику и мотивировать девочек! 💪",
                reply_markup=keyboard
            )
        else:
            keyboard = get_admin_keyboard()
            await update.message.reply_text(
                f"Привет, админ! 👑\n\n"
                f"Ты можешь управлять всей системой!",
                reply_markup=keyboard
            )
    else:
        await update.message.reply_text(
            f"Привет! 👋\n\n"
            f"Пожалуйста, напиши своё имя для регистрации.\n\n"
            f"📝 Доступные имена:\n"
            f"• Вика\n"
            f"• Юля\n"
            f"• Алла\n"
            f"• Аня\n"
            f"• ☀️ (солнышко)\n"
            f"• 🌸 (цветочек)\n\n"
            f"Напиши своё имя или эмодзи:"
        )

async def handle_video_creation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if not user_name:
        await update.message.reply_text("Пожалуйста, сначала зарегистрируйся с помощью /start")
        return ConversationHandler.END

    rates = db["users"][user_name]["rates"]
    video_types = [vtype for vtype in rates.keys() if vtype != "upload"]

    keyboard = []
    for vtype in video_types:
        keyboard.append([InlineKeyboardButton(f"{vtype.upper()} ({rates[vtype]} грн)", callback_data=f"video_type_{vtype}")])

    reply_markup = InlineKeyboardMarkup(keyboard)
    context.user_data['creating_video'] = True
    context.user_data['user_name'] = user_name

    await update.message.reply_text("Выбери тип видео:", reply_markup=reply_markup)
    return VIDEO_TYPE

async def video_type_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    video_type = query.data.replace("video_type_", "")
    context.user_data['video_type'] = video_type
    await query.edit_message_text(f"Отлично! Теперь напиши название видео:")
    return VIDEO_NAME

async def video_name_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    video_name = update.message.text
    user_name = context.user_data['user_name']
    video_type = context.user_data['video_type']
    amount = db["users"][user_name]["rates"][video_type]

    video_entry = {
        "date": datetime.now().isoformat(),
        "user": user_name,
        "type": video_type,
        "name": video_name,
        "amount": amount,
        "status": "created"
    }

    db["videos"].append(video_entry)
    db["balances"][user_name] += amount
    save_database(db)

    motivation = random.choice(MOTIVATIONAL_PHRASES["video_created"])
    motivation = motivation.format(name=user_name, video_type=video_type.upper(), amount=amount)
    keyboard = get_user_keyboard(user_name)
    await update.message.reply_text(motivation, reply_markup=keyboard)

    notification = f"🔔 {user_name} создала видео '{video_name}' ({video_type.upper()}) - +{amount} грн"
    for admin_id in [db["admins"]["admin"], db["admins"]["husband"]]:
        if admin_id:
            try:
                await context.bot.send_message(chat_id=admin_id, text=notification)
            except:
                pass

    context.user_data.clear()
    return ConversationHandler.END

async def handle_video_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if not user_name:
        await update.message.reply_text("Пожалуйста, сначала зарегистрируйся с помощью /start")
        return ConversationHandler.END

    context.user_data['uploading_video'] = True
    context.user_data['user_name'] = user_name
    await update.message.reply_text("Сколько видео ты загрузила? (напиши число)\nНапример: 3")
    return UPLOAD_COUNT

async def upload_count_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        count = int(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, напиши число!")
        return UPLOAD_COUNT

    user_name = context.user_data['user_name']
    upload_rate = db["users"][user_name]["rates"]["upload"]
    amount = count * upload_rate

    upload_entry = {
        "date": datetime.now().isoformat(),
        "user": user_name,
        "count": count,
        "amount": amount,
        "status": "uploaded"
    }

    db["uploads"].append(upload_entry)
    db["balances"][user_name] += amount
    save_database(db)

    motivation = random.choice(MOTIVATIONAL_PHRASES["video_uploaded"])
    motivation = motivation.format(name=user_name, count=count, amount=amount)
    keyboard = get_user_keyboard(user_name)
    await update.message.reply_text(motivation, reply_markup=keyboard)

    notification = f"🔔 {user_name} загрузила {count} видео - +{amount} грн"
    for admin_id in [db["admins"]["admin"], db["admins"]["husband"]]:
        if admin_id:
            try:
                await context.bot.send_message(chat_id=admin_id, text=notification)
            except:
                pass

    context.user_data.clear()
    return ConversationHandler.END

async def show_my_income(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if not user_name:
        await update.message.reply_text("Пожалуйста, сначала зарегистрируйся с помощью /start")
        return

    balance = db["balances"][user_name]
    videos_count = len([v for v in db["videos"] if v["user"] == user_name])
    uploads_count = sum([u["count"] for u in db["uploads"] if u["user"] == user_name])

    message = f"💰 Твой текущий доход: {balance} грн\n\n"
    message += f"📊 Статистика:\n"
    message += f"🎬 Создано видео: {videos_count}\n"
    message += f"📤 Загружено видео: {uploads_count}\n"

    await update.message.reply_text(message)

async def show_my_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if not user_name:
        await update.message.reply_text("Пожалуйста, сначала зарегистрируйся с помощью /start")
        return

    user_videos = [v for v in db["videos"] if v["user"] == user_name]
    user_uploads = [u for u in db["uploads"] if u["user"] == user_name]
    message = f"📊 Твоя статистика:\n\n"

    video_types = {}
    for video in user_videos:
        vtype = video["type"]
        if vtype not in video_types:
            video_types[vtype] = {"count": 0, "amount": 0}
        video_types[vtype]["count"] += 1
        video_types[vtype]["amount"] += video["amount"]

    for vtype, data in video_types.items():
        message += f"🎬 {vtype.upper()}: {data['count']} видео (+{data['amount']} грн)\n"

    if user_uploads:
        total_uploads = sum([u["count"] for u in user_uploads])
        total_upload_amount = sum([u["amount"] for u in user_uploads])
        message += f"📤 Загрузки: {total_uploads} видео (+{total_upload_amount} грн)\n"

    message += f"\n💰 Текущий баланс: {db['balances'][user_name]} грн"
    await update.message.reply_text(message)

async def show_full_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if not is_admin(telegram_id):
        await update.message.reply_text("У тебя нет прав для просмотра полной статистики")
        return

    message = "📊 ПОЛНАЯ СТАТИСТИКА КОМАНДЫ\n\n"

    for user_name in db["users"].keys():
        balance = db["balances"][user_name]
        videos_count = len([v for v in db["videos"] if v["user"] == user_name])
        uploads_count = sum([u["count"] for u in db["uploads"] if u["user"] == user_name])

        message += f"👤 {user_name}:\n"
        message += f" 💰 Баланс: {balance} грн\n"
        message += f" 🎬 Видео: {videos_count}\n"
        message += f" 📤 Загрузки: {uploads_count}\n\n"

    total_balance = sum(db["balances"].values())
    message += f"💎 ОБЩИЙ БАЛАНС: {total_balance} грн"
    await update.message.reply_text(message)

async def show_current_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if not is_admin(telegram_id):
        await update.message.reply_text("У тебя нет прав для просмотра баланса")
        return

    message = "⚙️ ТЕКУЩИЙ БАЛАНС (не выплачено):\n\n"
    for user_name, balance in db["balances"].items():
        message += f"👤 {user_name}: {balance} грн\n"

    total = sum(db["balances"].values())
    message += f"\n💎 ИТОГО: {total} грн"
    await update.message.reply_text(message)

async def show_all_videos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if not is_admin(telegram_id):
        await update.message.reply_text("У тебя нет прав для просмотра всех видео")
        return

    if not db["videos"] and not db["uploads"]:
        await update.message.reply_text("Пока нет созданных видео")
        return

    message = "🎬 ВСЕ ВИДЕО (не выплаченные):\n\n"

    for user_name in db["users"].keys():
        user_videos = [v for v in db["videos"] if v["user"] == user_name]
        user_uploads = [u for u in db["uploads"] if u["user"] == user_name]

        if not user_videos and not user_uploads:
            continue

        message += f"👤 {user_name}:\n"
        message += f"💰 Текущий баланс: {db['balances'][user_name]} грн\n\n"

        if user_videos:
            message += "   📹 СОЗДАННЫЕ ВИДЕО:\n"
            for video in user_videos:
                date = datetime.fromisoformat(video["date"]).strftime("%d.%m %H:%M")
                message += f"   • {date} | {video['type'].upper()} | \"{video['name']}\" | +{video['amount']} грн\n"
            message += "\n"

        if user_uploads:
            message += "   📤 ЗАГРУЖЕННЫЕ ВИДЕО:\n"
            for upload in user_uploads:
                date = datetime.fromisoformat(upload["date"]).strftime("%d.%m %H:%M")
                message += f"   • {date} | {upload['count']} видео на сайты | +{upload['amount']} грн\n"
            message += "\n"

        message += "─────────────────────\n\n"

    total_videos = len(db["videos"])
    total_uploads = sum([u["count"] for u in db["uploads"]])
    total_balance = sum(db["balances"].values())

    message += f"📊 ИТОГО:\n"
    message += f"🎬 Всего создано: {total_videos} видео\n"
    message += f"📤 Всего загружено: {total_uploads} видео\n"
    message += f"💎 Общий баланс: {total_balance} грн"

    await update.message.reply_text(message)

async def handle_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if telegram_id != db["admins"]["admin"]:
        await update.message.reply_text("Только админ может выплачивать зарплату")
        return ConversationHandler.END

    keyboard = []
    for user_name in db["users"].keys():
        balance = db["balances"][user_name]
        keyboard.append([InlineKeyboardButton(f"{user_name} ({balance} грн)", callback_data=f"pay_{user_name}")])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выбери кому выплатить зарплату:", reply_markup=reply_markup)
    return PAY_USER

async def payment_user_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    # Получаем имя пользователя из callback_data
    user_name = query.data.replace("pay_", "")
    
    context.user_data['payment_user'] = user_name
    balance = db["balances"][user_name]

    await query.edit_message_text(
        f"Выплата для {user_name}\n"
        f"Текущий баланс: {balance} грн\n\n"
        f"Напиши сумму для выплаты:"
    )
    return PAY_AMOUNT

async def payment_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = int(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, напиши число!")
        return PAY_AMOUNT

    user_name = context.user_data['payment_user']

    if amount > db["balances"][user_name]:
        await update.message.reply_text(
            f"Сумма превышает баланс {user_name} ({db['balances'][user_name]} грн)!\n"
            f"Напиши другую сумму:"
        )
        return PAY_AMOUNT

    # Архивируем видео и загрузки перед удалением
    user_videos = [v for v in db["videos"] if v["user"] == user_name]
    user_uploads = [u for u in db["uploads"] if u["user"] == user_name]

    payment_entry = {
        "date": datetime.now().isoformat(),
        "user": user_name,
        "amount": amount,
        "videos_paid": user_videos,
        "uploads_paid": user_uploads
    }

    db["payments"].append(payment_entry)
    db["balances"][user_name] -= amount

    db["videos"] = [v for v in db["videos"] if v["user"] != user_name]
    db["uploads"] = [u for u in db["uploads"] if u["user"] != user_name]

    save_database(db)

    user_telegram_id = db["users"][user_name].get("telegram_id")
    if user_telegram_id:
        motivation = random.choice(MOTIVATIONAL_PHRASES["payment"])
        motivation = motivation.format(name=user_name, amount=amount)
        try:
            await context.bot.send_message(chat_id=user_telegram_id, text=motivation)
        except:
            pass

    await update.message.reply_text(
        f"✅ Выплата выполнена!\n\n"
        f"👤 {user_name}\n"
        f"💰 Сумма: {amount} грн\n"
        f"📊 Новый баланс: {db['balances'][user_name]} грн\n"
        f"📹 Оплаченные видео архивированы!"
    )

    context.user_data.clear()
    return ConversationHandler.END

async def show_payment_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if not is_admin(telegram_id):
        await update.message.reply_text("У тебя нет прав для просмотра истории выплат")
        return

    if not db["payments"]:
        await update.message.reply_text("История выплат пуста")
        return

    message = "📈 ИСТОРИЯ ВЫПЛАТ:\n\n"
    recent_payments = db["payments"][-10:]

    for payment in reversed(recent_payments):
        date = datetime.fromisoformat(payment["date"]).strftime("%Y-%m-%d %H:%M")
        message += f"📅 {date}\n"
        message += f"👤 {payment['user']}: {payment['amount']} грн\n\n"

    total_paid = sum([p["amount"] for p in db["payments"]])
    message += f"💎 ВСЕГО ВЫПЛАЧЕНО: {total_paid} грн"
    await update.message.reply_text(message)

async def handle_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if telegram_id != db["admins"]["admin"]:
        await update.message.reply_text("Только админ может отправлять срочные сообщения")
        return ConversationHandler.END

    await update.message.reply_text("📢 Напиши срочное сообщение для всех девушек:")
    return BROADCAST_MSG

async def broadcast_message_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    broadcast_text = f"📢 СРОЧНО: {message}"

    sent_count = 0
    for user_name, config in db["users"].items():
        user_telegram_id = config.get("telegram_id")
        if user_telegram_id:
            try:
                await context.bot.send_message(chat_id=user_telegram_id, text=broadcast_text)
                sent_count += 1
            except:
                pass

    await update.message.reply_text(f"✅ Сообщение отправлено {sent_count} пользователям!")
    return ConversationHandler.END

async def send_husband_joke(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.effective_user.id

    if telegram_id != db["admins"]["husband"]:
        await update.message.reply_text("Только муж может отправлять прикольчики")
        return

    joke = random.choice(MOTIVATIONAL_PHRASES["husband_jokes"])

    sent_count = 0
    for user_name, config in db["users"].items():
        user_telegram_id = config.get("telegram_id")
        if user_telegram_id:
            try:
                await context.bot.send_message(chat_id=user_telegram_id, text=f"😂 {joke}")
                sent_count += 1
            except:
                pass

    await update.message.reply_text(f"✅ Прикольчик отправлен {sent_count} девочкам! 😄")

async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт полной статистики в Excel файл"""
    telegram_id = update.effective_user.id

    if telegram_id != db["admins"]["admin"]:
        await update.message.reply_text("Только админ может экспортировать статистику")
        return

    if not EXCEL_AVAILABLE:
        await update.message.reply_text(
            "❌ Библиотека openpyxl не установлена!\n"
            "Установите: pip install openpyxl"
        )
        return

    await update.message.reply_text("⏳ Создаю Excel файл... Подожди немного!")

    try:
        wb = Workbook()
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ЛИСТ 1: ТЕКУЩИЕ БАЛАНСЫ
        ws1 = wb.active
        ws1.title = "Текущие балансы"
        
        headers = ["Имя", "Текущий баланс (грн)", "Создано видео", "Загружено видео"]
        ws1.append(headers)
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for user_name in db["users"].keys():
            balance = db["balances"][user_name]
            videos_count = len([v for v in db["videos"] if v["user"] == user_name])
            uploads_count = sum([u["count"] for u in db["uploads"] if u["user"] == user_name])
            
            row = [user_name, balance, videos_count, uploads_count]
            ws1.append(row)
            
            for cell in ws1[ws1.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        total_balance = sum(db["balances"].values())
        total_videos = len(db["videos"])
        total_uploads = sum([u["count"] for u in db["uploads"]])
        
        ws1.append(["ИТОГО", total_balance, total_videos, total_uploads])
        for cell in ws1[ws1.max_row]:
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 18

        # ЛИСТ 2: ВСЕ ВИДЕО
        ws2 = wb.create_sheet("Все видео")
        
        headers = ["Дата", "Время", "Имя", "Тип видео", "Название", "Сумма (грн)"]
        ws2.append(headers)
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for video in db["videos"]:
            dt = datetime.fromisoformat(video["date"])
            row = [
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                video["user"],
                video["type"].upper(),
                video["name"],
                video["amount"]
            ]
            ws2.append(row)
            
            for cell in ws2[ws2.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 40
        ws2.column_dimensions['F'].width = 12

        # ЛИСТ 3: ВСЕ ЗАГРУЗКИ
        ws3 = wb.create_sheet("Все загрузки")
        
        headers = ["Дата", "Время", "Имя", "Количество видео", "Сумма (грн)"]
        ws3.append(headers)
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for upload in db["uploads"]:
            dt = datetime.fromisoformat(upload["date"])
            row = [
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                upload["user"],
                upload["count"],
                upload["amount"]
            ]
            ws3.append(row)
            
            for cell in ws3[ws3.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 10
        ws3.column_dimensions['C'].width = 12
        ws3.column_dimensions['D'].width = 18
        ws3.column_dimensions['E'].width = 12

        # ЛИСТ 4: ИСТОРИЯ ВЫПЛАТ
        ws4 = wb.create_sheet("История выплат")
        
        headers = ["Дата", "Время", "Имя", "Сумма выплаты (грн)", "Видео оплачено", "Загрузок оплачено"]
        ws4.append(headers)
        
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for payment in db["payments"]:
            dt = datetime.fromisoformat(payment["date"])
            videos_paid = len(payment.get("videos_paid", []))
            uploads_paid = sum([u["count"] for u in payment.get("uploads_paid", [])])
            
            row = [
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                payment["user"],
                payment["amount"],
                videos_paid,
                uploads_paid
            ]
            ws4.append(row)
            
            for cell in ws4[ws4.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        total_paid = sum([p["amount"] for p in db["payments"]])
        ws4.append(["", "", "ВСЕГО ВЫПЛАЧЕНО", total_paid, "", ""])
        for cell in ws4[ws4.max_row]:
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws4.column_dimensions['A'].width = 12
        ws4.column_dimensions['B'].width = 10
        ws4.column_dimensions['C'].width = 12
        ws4.column_dimensions['D'].width = 20
        ws4.column_dimensions['E'].width = 15
        ws4.column_dimensions['F'].width = 18

        # ЛИСТ 5: ТАРИФЫ
        ws5 = wb.create_sheet("Тарифы")
        
        headers = ["Имя", "A2E", "MAKEFILM", "GROK", "UPLOAD"]
        ws5.append(headers)
        
        for cell in ws5[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for user_name, config in db["users"].items():
            rates = config["rates"]
            row = [
                user_name,
                rates.get("a2e", "-"),
                rates.get("makefilm", "-"),
                rates.get("grok", "-"),
                rates.get("upload", "-")
            ]
            ws5.append(row)
            
            for cell in ws5[ws5.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws5.column_dimensions['A'].width = 12
        ws5.column_dimensions['B'].width = 12
        ws5.column_dimensions['C'].width = 15
        ws5.column_dimensions['D'].width = 12
        ws5.column_dimensions['E'].width = 12

        filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        with open(filename, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=filename,
                caption=f"📊 Полная статистика экспортирована!\n\n"
                        f"📅 Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                        f"💰 Общий баланс: {total_balance} грн\n"
                        f"💸 Всего выплачено: {total_paid} грн"
            )

        os.remove(filename)

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при создании Excel файла: {str(e)}")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    telegram_id = update.effective_user.id
    user_name = get_user_by_telegram_id(telegram_id)

    if not user_name and not is_admin(telegram_id):
        # Проверяем, есть ли это имя в списке пользователей
        if text in db["users"].keys():
            db["users"][text]["telegram_id"] = telegram_id
            save_database(db)
            keyboard = get_user_keyboard(text)
            await update.message.reply_text(
                f"✅ Регистрация успешна!\n\n"
                f"Привет, {text}! 👋\n"
                f"Теперь ты можешь использовать бота!",
                reply_markup=keyboard
            )
        else:
            await update.message.reply_text(
                f"❌ Имя '{text}' не найдено.\n\n"
                f"📝 Доступные имена:\n"
                f"• Вика\n"
                f"• Юля\n"
                f"• Алла\n"
                f"• Аня\n"
                f"• ☀️ (солнышко)\n"
                f"• 🌸 (цветочек)\n\n"
                f"Напиши своё имя или эмодзи:"
            )
        return

    if text == "🎬 Создала видео":
        await handle_video_creation(update, context)
    elif text == "📤 Загрузила видео":
        await handle_video_upload(update, context)
    elif text == "💰 Мой доход":
        await show_my_income(update, context)
    elif text == "📊 Моя статистика":
        await show_my_stats(update, context)
    elif text == "📊 Полная статистика":
        await show_full_stats(update, context)
    elif text == "⚙️ Текущий баланс":
        await show_current_balance(update, context)
    elif text == "💸 Выплатить зарплату":
        await handle_payment(update, context)
    elif text == "📈 История выплат":
        await show_payment_history(update, context)
    elif text == "🎬 Все видео":
        await show_all_videos(update, context)
    elif text == "📢 Срочное сообщение всем":
        await handle_broadcast(update, context)
    elif text == "😂 Отправить прикольчик" or text == "💪 Мотивация девочкам":
        await send_husband_joke(update, context)
    elif text == "📊 Экспорт в Excel":
        await export_to_excel(update, context)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Операция отменена")
    return ConversationHandler.END

def main():
    TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

    if not TOKEN:
        print("❌ Ошибка: Не найден TELEGRAM_BOT_TOKEN")
        print("Установите токен: export TELEGRAM_BOT_TOKEN='your_token_here'")
        return

    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))

    video_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^🎬 Создала видео$"), handle_video_creation)],
        states={
            VIDEO_TYPE: [CallbackQueryHandler(video_type_selected, pattern="^video_type_")],
            VIDEO_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, video_name_entered)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    upload_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📤 Загрузила видео$"), handle_video_upload)],
        states={
            UPLOAD_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, upload_count_entered)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    payment_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^💸 Выплатить зарплату$"), handle_payment)],
        states={
            PAY_USER: [CallbackQueryHandler(payment_user_selected, pattern="^pay_")],
            PAY_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_amount_entered)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    broadcast_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📢 Срочное сообщение всем$"), handle_broadcast)],
        states={
            BROADCAST_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_message_entered)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    application.add_handler(video_conv_handler)
    application.add_handler(upload_conv_handler)
    application.add_handler(payment_conv_handler)
    application.add_handler(broadcast_conv_handler)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("🚀 Бот запущен!")
    print(f"📊 Excel экспорт: {'✅ Доступен' if EXCEL_AVAILABLE else '❌ Требуется установка openpyxl'}")
    print(f"👥 Пользователей в системе: {len(db['users'])}")
    print(f"   • Вика, Юля, Алла, Аня, ☀️, 🌸")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
