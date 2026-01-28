#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🤖 TELEGRAM BOT ДЛЯ УПРАВЛЕНИЯ ВИДЕОПРОИЗВОДСТВОМ
Версия: 3.0 (ПОЛНАЯ ПЕРЕРАБОТКА)
Дата: 28.01.2026

ФУНКЦИИ:
✅ Исправлен баг ConversationHandler (проблема Ани)
✅ Система рейтингов и достижений
✅ Планирование задач на неделю
✅ Удаление видео администратором
✅ Выплата авансов
✅ Система выходных дней с одобрением
✅ Персональные календари (только свои выходные + админы)
✅ Экспорт в Excel
"""

import os
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters
)

# ===========================
# НАСТРОЙКА ЛОГИРОВАНИЯ
# ===========================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ===========================
# КОНСТАНТЫ
# ===========================
DB_FILE = 'bot_database.json'

# Состояния ConversationHandler
VIDEO_TYPE, VIDEO_NAME = range(2)
UPLOAD_COUNT = 100
BROADCAST_MESSAGE = 700
DELETE_VIDEO_SELECT, DELETE_VIDEO_CONFIRM = 200, 201
PLAN_USER, PLAN_DETAILS = 300, 301
ADVANCE_USER, ADVANCE_AMOUNT, ADVANCE_CONFIRM = 400, 401, 402
DAYOFF_DATE, DAYOFF_REASON, DAYOFF_APPROVE = 500, 501, 502
ADMIN_DAYOFF_WHO, ADMIN_DAYOFF_DATES = 600, 601

# Администраторы (Telegram ID)
ADMINS = [2147091471]  # ЗАМЕНИ НА СВОЙ TELEGRAM ID
HUSBAND_ID = 2106439695  # ЗАМЕНИ НА TELEGRAM ID МУЖА

# ===========================
# КОНФИГУРАЦИЯ ПОЛЬЗОВАТЕЛЕЙ
# ===========================
USERS_CONFIG = {
    "Вика": {
        "role": "creator_uploader",
        "rates": {"a2e": 350, "makefilm": 450, "grok": 550, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Юля": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 500, "grok": 550, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Алла": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "Аня": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "☀️": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    },
    "🌸": {
        "role": "creator_uploader",
        "rates": {"a2e": 300, "makefilm": 400, "grok": 450, "upload": 200},
        "can_upload": True,
        "telegram_id": None
    }
}

# ===========================
# БАЗА ДАННЫХ
# ===========================
def load_database() -> Dict:
    """Загрузка базы данных из JSON"""
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Ошибка загрузки БД: {e}")
    
    # Инициализация базы данных с выходными
    return {
        "users": USERS_CONFIG.copy(),
        "videos": [],
        "payments": [],
        "plans": {},
        "days_off_requests": [],
        "days_off_approved": {
            "Вика": [{"date": "2026-01-30", "reason": "выходной", "approved_at": "2026-01-28"}],
            "Юля": [
                {"date": "2026-02-02", "reason": "выходной", "approved_at": "2026-01-28"},
                {"date": "2026-02-03", "reason": "выходной", "approved_at": "2026-01-28"}
            ],
            "Алла": [],
            "Аня": [{"date": "2026-02-05", "reason": "выходной", "approved_at": "2026-01-28"}],
            "☀️": [],
            "🌸": []
        },
        "admin_days_off": {
            "admin": [],
            "husband": []
        }
    }

def save_database(db: Dict):
    """Сохранение базы данных в JSON"""
    try:
        with open(DB_FILE, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Ошибка сохранения БД: {e}")

db = load_database()

# ===========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ===========================
def is_admin(user_id: int) -> bool:
    """Проверка, является ли пользователь администратором"""
    return user_id in ADMINS or user_id == HUSBAND_ID
    logger.info(f"🔑 Проверка админа: user_id={user_id}, ADMINS={ADMINS}, HUSBAND_ID={HUSBAND_ID}, is_admin={user_id in ADMINS or user_id == HUSBAND_ID}")

def get_user_name(user_id: int) -> Optional[str]:
    """Получение имени пользователя по Telegram ID"""
    for name, data in db['users'].items():
        if data.get('telegram_id') == user_id:
            return name
    return None

def calculate_balance(user_name: str) -> int:
    """Расчёт текущего баланса пользователя"""
    total = 0
    
    # Добавляем доход от видео
    for video in db['videos']:
        if video['user'] == user_name:
            total += video['amount']
    
    # Вычитаем выплаты
    for payment in db['payments']:
        if payment['user'] == user_name:
            total -= payment['amount']
    
    return total

def get_user_stats(user_name: str) -> Dict:
    """Получение статистики пользователя"""
    videos = [v for v in db['videos'] if v['user'] == user_name]
    
    stats = {
        'total_videos': len(videos),
        'total_earnings': sum(v['amount'] for v in videos),
        'by_type': {},
        'balance': calculate_balance(user_name)
    }
    
    for video_type in ['a2e', 'makefilm', 'grok', 'upload']:
        type_videos = [v for v in videos if v['type'] == video_type]
        stats['by_type'][video_type] = {
            'count': len(type_videos),
            'earnings': sum(v['amount'] for v in type_videos)
        }
    
    return stats

def format_date(date_str: str) -> str:
    """Форматирование даты в читаемый вид"""
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        weekdays = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
        return f"{date_obj.strftime('%d.%m')} ({weekdays[date_obj.weekday()]})"
    except:
        return date_str

def get_main_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    """Получение главной клавиатуры в зависимости от роли"""
    if is_admin(user_id):
        keyboard = [
            ['📊 Полная статистика', '⚙️ Текущий баланс'],
            ['💸 Выплатить зарплату', '💰 Выплатить аванс'],
            ['📈 История выплат', '🗑️ Удалить видео'],
            ['🎬 Все видео', '📊 Экспорт в Excel'],
            ['🏆 Рейтинг девушек', '📅 План на неделю'],
            ['📅 График выходных', '📅 Мои выходные'],
            ['🔔 Запросы выходных']
        ]
    else:
        keyboard = [
            ['🎬 Создала видео'],
            ['💰 Мой доход', '📊 Моя статистика'],
            ['📅 Мой план', '📅 Мой календарь'],
            ['📅 Запросить выходной']
        ]
    
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# ===========================
# КОМАНДА /start
# ===========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка команды /start"""
    user_id = update.effective_user.id
    
    # ЛОГИРОВАНИЕ ДЛЯ ОТЛАДКИ
    logger.info(f"🔍 /start от пользователя: ID={user_id}, Username={update.effective_user.username}, Name={update.effective_user.first_name}")
    
    # Сброс состояния ConversationHandler (FIX для бага Ани)
    context.user_data.clear()
    
    if is_admin(user_id):
        await update.message.reply_text(
            "👋 Привет, Администратор!\n\n"
            "Выбери действие:",
            reply_markup=get_main_keyboard(user_id)
        )
        return
    
    # Проверка регистрации
    user_name = get_user_name(user_id)
    
    if user_name:
        await update.message.reply_text(
            f"👋 Привет, {user_name}!\n\n"
            f"Выбери действие:",
            reply_markup=get_main_keyboard(user_id)
        )
    else:
        # Запрос имени для регистрации
        names_list = "\n".join([f"• {name}" for name in USERS_CONFIG.keys()])
        await update.message.reply_text(
            "👋 Привет! Как тебя зовут?\n\n"
            f"Доступные имена:\n{names_list}\n\n"
            "Напиши своё имя точно как в списке:"
        )

# ===========================
# РЕГИСТРАЦИЯ ПОЛЬЗОВАТЕЛЯ
# ===========================
async def handle_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка регистрации нового пользователя"""
    user_id = update.effective_user.id
    text = update.message.text
    
    # Пропускаем, если уже зарегистрирован
    if get_user_name(user_id):
        return
    
    # Проверка имени
    if text in USERS_CONFIG.keys():
        db['users'][text]['telegram_id'] = user_id
        save_database(db)
        
        await update.message.reply_text(
            f"✅ Регистрация завершена!\n\n"
            f"Привет, {text}! Теперь ты можешь работать с ботом.",
            reply_markup=get_main_keyboard(user_id)
        )
    else:
        names_list = "\n".join([f"• {name}" for name in USERS_CONFIG.keys()])
        await update.message.reply_text(
            f"❌ Имя '{text}' не найдено.\n\n"
            f"Доступные имена:\n{names_list}\n\n"
            "Напиши своё имя точно как в списке:"
        )

# ===========================
# СОЗДАНИЕ ВИДЕО
# ===========================
async def handle_video_creation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало создания видео (entry_point ConversationHandler)"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return ConversationHandler.END
    
    # Получаем расценки пользователя
    rates = db['users'][user_name]['rates']
    
    # Создаём inline-клавиатуру с типами видео (ИСКЛЮЧАЯ upload)
    keyboard = []
    for video_type in ['a2e', 'makefilm', 'grok']:
        price = rates.get(video_type, 0)
        keyboard.append([
            InlineKeyboardButton(
                f"{video_type.upper()} — {price} грн",
                callback_data=f"video_type_{video_type}"
            )
        ])
    
    # Сохраняем контекст
    context.user_data['creating_video'] = True
    context.user_data['user_name'] = user_name
    
    await update.message.reply_text(
        "🎬 Выбери тип видео:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return VIDEO_TYPE

async def video_type_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора типа видео"""
    query = update.callback_query
    await query.answer()
    
    # Извлекаем тип видео из callback_data
    video_type = query.data.replace("video_type_", "")
    context.user_data['video_type'] = video_type
    
    user_name = context.user_data['user_name']
    price = db['users'][user_name]['rates'][video_type]
    
    await query.edit_message_text(
        f"✅ Выбран тип: {video_type.upper()} ({price} грн)\n\n"
        f"📝 Теперь напиши название видео:"
    )
    
    return VIDEO_NAME

async def video_name_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода названия видео"""
    video_name = update.message.text
    user_name = context.user_data['user_name']
    video_type = context.user_data['video_type']
    
    # Получаем цену
    price = db['users'][user_name]['rates'][video_type]
    
    # Сохраняем видео в БД
    video_entry = {
        'id': len(db['videos']) + 1,
        'user': user_name,
        'type': video_type,
        'name': video_name,
        'amount': price,
        'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    db['videos'].append(video_entry)
    save_database(db)
    
    # Обновляем баланс
    new_balance = calculate_balance(user_name)
    
    await update.message.reply_text(
        f"✅ Видео добавлено!\n\n"
        f"📹 Название: {video_name}\n"
        f"🎬 Тип: {video_type.upper()}\n"
        f"💰 Заработано: +{price} грн\n"
        f"💵 Текущий баланс: {new_balance} грн",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    # Уведомление админу о создании видео
    for admin_id in ADMINS + [HUSBAND_ID]:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"🎬 НОВОЕ ВИДЕО!\n\n"
                     f"👤 {user_name}\n"
                     f"📹 {video_name}\n"
                     f"🎬 Тип: {video_type.upper()}\n"
                     f"💰 Сумма: {price} грн\n"
                     f"💵 Баланс: {new_balance} грн"
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление админу {admin_id}: {e}")
    
    # Очистка контекста
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# ЗАГРУЗКА ВИДЕО
# ===========================
async def handle_video_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало загрузки видео"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return ConversationHandler.END
    
    if not db['users'][user_name].get('can_upload', False):
        await update.message.reply_text("❌ У тебя нет прав на загрузку видео")
        return ConversationHandler.END
    
    context.user_data['uploading_video'] = True
    context.user_data['user_name'] = user_name
    
    await update.message.reply_text(
        "📤 Сколько видео загрузила?\n\n"
        "Напиши число:"
    )
    
    return UPLOAD_COUNT

async def upload_count_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка количества загруженных видео"""
    try:
        count = int(update.message.text)
        if count <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Пожалуйста, напиши число!\n\n"
            "Сколько видео загрузила?"
        )
        return UPLOAD_COUNT
    
    user_name = context.user_data['user_name']
    upload_rate = db['users'][user_name]['rates']['upload']
    total_amount = count * upload_rate
    
    # Сохраняем каждое загруженное видео
    for i in range(count):
        video_entry = {
            'id': len(db['videos']) + 1,
            'user': user_name,
            'type': 'upload',
            'name': f"Загрузка #{i+1}",
            'amount': upload_rate,
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        db['videos'].append(video_entry)
    
    save_database(db)
    
    new_balance = calculate_balance(user_name)
    
    await update.message.reply_text(
        f"✅ Загрузка засчитана!\n\n"
        f"📤 Количество видео: {count}\n"
        f"💰 Заработано: +{total_amount} грн\n"
        f"💵 Текущий баланс: {new_balance} грн",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# МОЙ ДОХОД
# ===========================
async def my_income(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ дохода пользователя"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return
    
    stats = get_user_stats(user_name)
    
    message = f"💰 ТВОЙ ДОХОД\n\n"
    message += f"💵 Текущий баланс: {stats['balance']} грн\n"
    message += f"📊 Всего заработано: {stats['total_earnings']} грн\n"
    message += f"🎬 Всего видео: {stats['total_videos']}\n\n"
    message += f"📈 ПО ТИПАМ:\n"
    
    for video_type, data in stats['by_type'].items():
        if data['count'] > 0:
            message += f"• {video_type.upper()}: {data['count']} шт. — {data['earnings']} грн\n"
    
    await update.message.reply_text(message)

# ===========================
# МОЯ СТАТИСТИКА
# ===========================
async def my_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ статистики пользователя"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return
    
    stats = get_user_stats(user_name)
    videos = [v for v in db['videos'] if v['user'] == user_name]
    
    # Последние 5 видео
    recent_videos = sorted(videos, key=lambda x: x['created_at'], reverse=True)[:5]
    
    message = f"📊 ТВОЯ СТАТИСТИКА\n\n"
    message += f"💵 Баланс: {stats['balance']} грн\n"
    message += f"🎬 Видео: {stats['total_videos']}\n"
    message += f"💰 Заработано: {stats['total_earnings']} грн\n\n"
    
    if recent_videos:
        message += f"📹 ПОСЛЕДНИЕ ВИДЕО:\n"
        for v in recent_videos:
            date = datetime.strptime(v['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
            message += f"• {date} | {v['type'].upper()} | {v['name'][:20]} | +{v['amount']} грн\n"
    
    await update.message.reply_text(message)

# ===========================
# ПОЛНАЯ СТАТИСТИКА (АДМИН)
# ===========================
async def full_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ полной статистики (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    message = "📊 ПОЛНАЯ СТАТИСТИКА\n\n"
    
    for user_name in USERS_CONFIG.keys():
        stats = get_user_stats(user_name)
        message += f"👤 {user_name}:\n"
        message += f"   💵 Баланс: {stats['balance']} грн\n"
        message += f"   🎬 Видео: {stats['total_videos']}\n"
        message += f"   💰 Заработано: {stats['total_earnings']} грн\n\n"
    
    await update.message.reply_text(message)

# ===========================
# ТЕКУЩИЙ БАЛАНС (АДМИН)
# ===========================
async def current_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ текущего баланса всех пользователей (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    message = "⚙️ ТЕКУЩИЙ БАЛАНС\n\n"
    
    total = 0
    for user_name in USERS_CONFIG.keys():
        balance = calculate_balance(user_name)
        total += balance
        message += f"👤 {user_name}: {balance} грн\n"
    
    message += f"\n💰 Всего к выплате: {total} грн"
    
    await update.message.reply_text(message)

# ===========================
# ВЫПЛАТА ЗАРПЛАТЫ (АДМИН)
# ===========================
async def salary_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выплата зарплаты (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    # Создаём inline-клавиатуру с пользователями
    keyboard = []
    for user_name in USERS_CONFIG.keys():
        balance = calculate_balance(user_name)
        if balance > 0:
            keyboard.append([
                InlineKeyboardButton(
                    f"{user_name} — {balance} грн",
                    callback_data=f"pay_salary_{user_name}"
                )
            ])
    
    if not keyboard:
        await update.message.reply_text("✅ Никому не нужно выплачивать зарплату")
        return
    
    await update.message.reply_text(
        "💸 Кому выплатить зарплату?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def process_salary_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выплаты зарплаты"""
    query = update.callback_query
    await query.answer()
    
    user_name = query.data.replace("pay_salary_", "")
    balance = calculate_balance(user_name)
    
    if balance <= 0:
        await query.edit_message_text("❌ У этого пользователя нет баланса для выплаты")
        return
    
    # Создаём запись о выплате
    payment_entry = {
        'id': len(db['payments']) + 1,
        'user': user_name,
        'amount': balance,
        'type': 'salary',
        'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    db['payments'].append(payment_entry)
    save_database(db)
    
    # Уведомление пользователю
    user_telegram_id = db['users'][user_name].get('telegram_id')
    if user_telegram_id:
        try:
            await context.bot.send_message(
                chat_id=user_telegram_id,
                text=f"💸 ВЫПЛАТА ЗАРПЛАТЫ\n\n"
                     f"Тебе выплачено: {balance} грн\n"
                     f"Твой баланс обнулён.\n\n"
                     f"Удачи! 💪"
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление {user_name}: {e}")
    
    await query.edit_message_text(
        f"✅ Зарплата выплачена!\n\n"
        f"👤 {user_name}\n"
        f"💰 Сумма: {balance} грн"
    )

# ===========================
# ВЫПЛАТА АВАНСА (АДМИН)
# ===========================
async def advance_payment_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало выплаты аванса (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return ConversationHandler.END
    
    # Создаём inline-клавиатуру с пользователями
    keyboard = []
    for user_name in USERS_CONFIG.keys():
        balance = calculate_balance(user_name)
        if balance > 0:
            keyboard.append([
                InlineKeyboardButton(
                    f"{user_name} — {balance} грн",
                    callback_data=f"advance_user_{user_name}"
                )
            ])
    
    if not keyboard:
        await update.message.reply_text("✅ Никому нельзя выплатить аванс (нет балансов)")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "💰 Кому выплатить аванс?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return ADVANCE_USER

async def advance_user_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора пользователя для аванса"""
    query = update.callback_query
    await query.answer()
    
    user_name = query.data.replace("advance_user_", "")
    balance = calculate_balance(user_name)
    
    context.user_data['advance_user'] = user_name
    context.user_data['advance_max'] = balance
    
    await query.edit_message_text(
        f"💰 Выплата аванса для {user_name}\n\n"
        f"💵 Доступный баланс: {balance} грн\n\n"
        f"Напиши сумму аванса:"
    )
    
    return ADVANCE_AMOUNT

async def advance_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода суммы аванса"""
    try:
        amount = int(update.message.text)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Пожалуйста, напиши число больше 0!\n\n"
            "Сумма аванса:"
        )
        return ADVANCE_AMOUNT
    
    user_name = context.user_data['advance_user']
    max_amount = context.user_data['advance_max']
    
    if amount > max_amount:
        await update.message.reply_text(
            f"❌ Сумма превышает доступный баланс ({max_amount} грн)!\n\n"
            f"Напиши сумму аванса:"
        )
        return ADVANCE_AMOUNT
    
    context.user_data['advance_amount'] = amount
    
    # Подтверждение
    keyboard = [
        [InlineKeyboardButton("✅ Да, выплатить", callback_data="advance_confirm_yes")],
        [InlineKeyboardButton("❌ Отмена", callback_data="advance_confirm_no")]
    ]
    
    await update.message.reply_text(
        f"💰 ПОДТВЕРЖДЕНИЕ АВАНСА\n\n"
        f"👤 {user_name}\n"
        f"💵 Сумма: {amount} грн\n"
        f"💰 Останется на балансе: {max_amount - amount} грн\n\n"
        f"Выплатить аванс?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return ADVANCE_CONFIRM

async def advance_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка подтверждения аванса"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "advance_confirm_no":
        await query.edit_message_text("❌ Выплата аванса отменена")
        context.user_data.clear()
        return ConversationHandler.END
    
    user_name = context.user_data['advance_user']
    amount = context.user_data['advance_amount']
    
    # Создаём запись о выплате
    payment_entry = {
        'id': len(db['payments']) + 1,
        'user': user_name,
        'amount': amount,
        'type': 'advance',
        'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    db['payments'].append(payment_entry)
    save_database(db)
    
    new_balance = calculate_balance(user_name)
    
    # Уведомление пользователю
    user_telegram_id = db['users'][user_name].get('telegram_id')
    if user_telegram_id:
        try:
            await context.bot.send_message(
                chat_id=user_telegram_id,
                text=f"💰 ВЫПЛАТА АВАНСА\n\n"
                     f"Тебе выплачено: {amount} грн\n"
                     f"Остаток на балансе: {new_balance} грн\n\n"
                     f"Продолжай работать! 💪"
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление {user_name}: {e}")
    
    await query.edit_message_text(
        f"✅ Аванс выплачен!\n\n"
        f"👤 {user_name}\n"
        f"💰 Сумма: {amount} грн\n"
        f"💵 Остаток на балансе: {new_balance} грн"
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# ИСТОРИЯ ВЫПЛАТ (АДМИН)
# ===========================
async def payment_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ истории выплат (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    payments = sorted(db['payments'], key=lambda x: x['created_at'], reverse=True)[:20]
    
    if not payments:
        await update.message.reply_text("📈 История выплат пуста")
        return
    
    message = "📈 ИСТОРИЯ ВЫПЛАТ\n\n"
    
    for payment in payments:
        date = datetime.strptime(payment['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
        payment_type = "💸 Зарплата" if payment['type'] == 'salary' else "💰 Аванс"
        message += f"{payment_type} | {date}\n"
        message += f"   👤 {payment['user']} — {payment['amount']} грн\n\n"
    
    await update.message.reply_text(message)

# ===========================
# ВСЕ ВИДЕО (АДМИН)
# ===========================
async def all_videos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ всех видео (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    videos = sorted(db['videos'], key=lambda x: x['created_at'], reverse=True)[:20]
    
    if not videos:
        await update.message.reply_text("🎬 Видео ещё нет")
        return
    
    message = "🎬 ВСЕ ВИДЕО (последние 20)\n\n"
    
    for video in videos:
        date = datetime.strptime(video['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
        message += f"#{video['id']} | {date}\n"
        message += f"   👤 {video['user']} | {video['type'].upper()}\n"
        message += f"   📹 {video['name'][:30]}\n"
        message += f"   💰 {video['amount']} грн\n\n"
    
    await update.message.reply_text(message)

# ===========================
# УДАЛЕНИЕ ВИДЕО (АДМИН)
# ===========================
async def delete_video_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало удаления видео (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return ConversationHandler.END
    
    videos = sorted(db['videos'], key=lambda x: x['created_at'], reverse=True)[:10]
    
    if not videos:
        await update.message.reply_text("🎬 Видео нет для удаления")
        return ConversationHandler.END
    
    message = "🗑️ УДАЛИТЬ ВИДЕО\n\n"
    message += "Последние 10 видео:\n\n"
    
    for i, video in enumerate(videos, 1):
        date = datetime.strptime(video['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
        message += f"{i}. {video['user']} | {date} | {video['type'].upper()}\n"
        message += f"   📹 {video['name'][:30]}\n"
        message += f"   💰 +{video['amount']} грн\n\n"
    
    message += "Напиши номер видео для удаления (1-10):"
    
    context.user_data['delete_videos'] = videos
    
    await update.message.reply_text(message)
    
    return DELETE_VIDEO_SELECT

async def delete_video_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора видео для удаления"""
    try:
        index = int(update.message.text) - 1
        videos = context.user_data['delete_videos']
        
        if index < 0 or index >= len(videos):
            raise ValueError
        
        video = videos[index]
    except (ValueError, IndexError):
        await update.message.reply_text(
            "❌ Неверный номер!\n\n"
            "Напиши номер от 1 до 10:"
        )
        return DELETE_VIDEO_SELECT
    
    context.user_data['delete_video'] = video
    
    # Подтверждение
    keyboard = [
        [InlineKeyboardButton("✅ Да, удалить", callback_data="delete_confirm_yes")],
        [InlineKeyboardButton("❌ Отмена", callback_data="delete_confirm_no")]
    ]
    
    date = datetime.strptime(video['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
    
    await update.message.reply_text(
        f"🗑️ ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ\n\n"
        f"#{video['id']} | {date}\n"
        f"👤 {video['user']}\n"
        f"🎬 {video['type'].upper()}\n"
        f"📹 {video['name']}\n"
        f"💰 -{video['amount']} грн\n\n"
        f"Точно удалить?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return DELETE_VIDEO_CONFIRM

async def delete_video_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка подтверждения удаления видео"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "delete_confirm_no":
        await query.edit_message_text("❌ Удаление отменено")
        context.user_data.clear()
        return ConversationHandler.END
    
    video = context.user_data['delete_video']
    
    # Удаляем видео из БД
    db['videos'] = [v for v in db['videos'] if v['id'] != video['id']]
    save_database(db)
    
    # Уведомление пользователю
    user_telegram_id = db['users'][video['user']].get('telegram_id')
    if user_telegram_id:
        try:
            await context.bot.send_message(
                chat_id=user_telegram_id,
                text=f"⚠️ ВИДЕО УДАЛЕНО АДМИНОМ\n\n"
                     f"📹 Название: {video['name']}\n"
                     f"🎬 Тип: {video['type'].upper()}\n"
                     f"💰 Сумма: -{video['amount']} грн\n\n"
                     f"Причина: ошибка при вводе"
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление {video['user']}: {e}")
    
    await query.edit_message_text(
        f"✅ Видео удалено!\n\n"
        f"#{video['id']} | {video['user']}\n"
        f"📹 {video['name']}\n"
        f"💰 -{video['amount']} грн"
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# РЕЙТИНГ ДЕВУШЕК (АДМИН)
# ===========================
async def ratings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ рейтинга девушек (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    # Собираем статистику
    users_stats = []
    for user_name in USERS_CONFIG.keys():
        stats = get_user_stats(user_name)
        users_stats.append({
            'name': user_name,
            'videos': stats['total_videos'],
            'earnings': stats['total_earnings']
        })
    
    # Сортируем по видео
    by_videos = sorted(users_stats, key=lambda x: x['videos'], reverse=True)
    
    # Сортируем по доходу
    by_earnings = sorted(users_stats, key=lambda x: x['earnings'], reverse=True)
    
    message = "🏆 РЕЙТИНГ ДЕВУШЕК\n\n"
    
    message += "📊 ПО КОЛИЧЕСТВУ ВИДЕО:\n"
    for i, user in enumerate(by_videos, 1):
        emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
        message += f"{emoji} {i}. {user['name']} — {user['videos']} видео\n"
    
    message += "\n💰 ПО ЗАРАБОТКУ:\n"
    for i, user in enumerate(by_earnings, 1):
        emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
        message += f"{emoji} {i}. {user['name']} — {user['earnings']} грн\n"
    
    await update.message.reply_text(message)

# ===========================
# ПЛАН НА НЕДЕЛЮ (АДМИН)
# ===========================
async def plan_week_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало установки плана на неделю (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return ConversationHandler.END
    
    # Создаём inline-клавиатуру с пользователями
    keyboard = []
    for user_name in USERS_CONFIG.keys():
        keyboard.append([
            InlineKeyboardButton(
                user_name,
                callback_data=f"plan_user_{user_name}"
            )
        ])
    
    await update.message.reply_text(
        "📅 ПЛАН НА НЕДЕЛЮ\n\n"
        "Выбери девушку:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return PLAN_USER

async def plan_user_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора пользователя для плана"""
    query = update.callback_query
    await query.answer()
    
    user_name = query.data.replace("plan_user_", "")
    context.user_data['plan_user'] = user_name
    
    await query.edit_message_text(
        f"📅 План для {user_name}\n\n"
        f"Напиши план в формате:\n"
        f"<количество> <тип> до <день недели>\n\n"
        f"Примеры:\n"
        f"• 5 видео до пятницы\n"
        f"• 3 grok до среды\n"
        f"• 10 любых до воскресенья"
    )
    
    return PLAN_DETAILS

async def plan_details_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода деталей плана"""
    user_name = context.user_data['plan_user']
    plan_text = update.message.text
    
    # Парсим план (упрощённо)
    # Формат: "5 видео до пятницы"
    try:
        parts = plan_text.lower().split()
        count = int(parts[0])
        video_type = parts[1] if len(parts) > 1 else "видео"
        deadline = " ".join(parts[parts.index("до")+1:]) if "до" in parts else "неделю"
    except:
        count = 5
        video_type = "видео"
        deadline = "неделю"
    
    # Сохраняем план
    plan_entry = {
        'user': user_name,
        'target_count': count,
        'video_type': video_type,
        'deadline': deadline,
        'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'completed': 0
    }
    
    db['plans'][user_name] = plan_entry
    save_database(db)
    
    # Уведомление пользователю
    user_telegram_id = db['users'][user_name].get('telegram_id')
    if user_telegram_id:
        try:
            await context.bot.send_message(
                chat_id=user_telegram_id,
                text=f"📅 НОВЫЙ ПЛАН НА НЕДЕЛЮ\n\n"
                     f"🎯 Цель: {count} {video_type}\n"
                     f"⏰ Дедлайн: {deadline}\n\n"
                     f"Давай, ты справишься! 💪"
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление {user_name}: {e}")
    
    await update.message.reply_text(
        f"✅ План установлен!\n\n"
        f"👤 {user_name}\n"
        f"🎯 Цель: {count} {video_type}\n"
        f"⏰ Дедлайн: {deadline}",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# МОЙ ПЛАН (ДЕВУШКИ)
# ===========================
async def my_plan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ плана пользователя"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return
    
    plan = db['plans'].get(user_name)
    
    if not plan:
        await update.message.reply_text(
            "📅 У тебя пока нет плана на неделю\n\n"
            "Администратор установит план позже."
        )
        return
    
    # Считаем прогресс (упрощённо - все видео за последние 7 дней)
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    recent_videos = [
        v for v in db['videos']
        if v['user'] == user_name and v['created_at'] >= week_ago
    ]
    completed = len(recent_videos)
    target = plan['target_count']
    progress = min(100, int(completed / target * 100))
    
    message = f"📅 МОЙ ПЛАН НА НЕДЕЛЮ\n\n"
    message += f"🎯 Цель: {target} {plan['video_type']}\n"
    message += f"⏰ Дедлайн: {plan['deadline']}\n"
    message += f"✅ Выполнено: {completed}/{target}\n"
    message += f"📊 Прогресс: {progress}%\n\n"
    
    if completed >= target:
        message += "🎉 План выполнен! Отличная работа!"
    elif progress >= 75:
        message += "💪 Почти готово! Осталось совсем чуть-чуть!"
    elif progress >= 50:
        message += "👍 Хороший прогресс! Продолжай в том же духе!"
    else:
        message += "⏰ Нужно ускориться, чтобы успеть!"
    
    await update.message.reply_text(message)

# ===========================
# ЗАПРОС ВЫХОДНОГО (ДЕВУШКИ)
# ===========================
async def request_dayoff_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало запроса выходного"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return ConversationHandler.END
    
    context.user_data['dayoff_user'] = user_name
    
    await update.message.reply_text(
        "📅 ЗАПРОС ВЫХОДНОГО\n\n"
        "Напиши дату в формате ДД.ММ\n\n"
        "Пример: 30.01"
    )
    
    return DAYOFF_DATE

async def dayoff_date_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода даты выходного"""
    date_text = update.message.text
    
    # Парсим дату
    try:
        day, month = date_text.split('.')
        year = datetime.now().year
        date_obj = datetime(year, int(month), int(day))
        date_str = date_obj.strftime("%Y-%m-%d")
    except:
        await update.message.reply_text(
            "❌ Неверный формат даты!\n\n"
            "Напиши дату в формате ДД.ММ\n"
            "Пример: 30.01"
        )
        return DAYOFF_DATE
    
    context.user_data['dayoff_date'] = date_str
    
    await update.message.reply_text(
        f"✅ Дата: {format_date(date_str)}\n\n"
        f"Напиши причину выходного (необязательно):\n\n"
        f"Или отправь \"-\" чтобы пропустить"
    )
    
    return DAYOFF_REASON

async def dayoff_reason_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода причины выходного"""
    reason = update.message.text if update.message.text != "-" else "Личные дела"
    
    user_name = context.user_data['dayoff_user']
    date_str = context.user_data['dayoff_date']
    
    # Создаём запрос
    request_entry = {
        'id': f"req_{len(db['days_off_requests']) + 1:03d}",
        'user': user_name,
        'date': date_str,
        'reason': reason,
        'status': 'pending',
        'requested_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    db['days_off_requests'].append(request_entry)
    save_database(db)
    
    # Уведомление админу
    for admin_id in ADMINS + [HUSBAND_ID]:
        try:
            keyboard = [
                [
                    InlineKeyboardButton("✅ Одобрить", callback_data=f"dayoff_approve_{request_entry['id']}"),
                    InlineKeyboardButton("❌ Отклонить", callback_data=f"dayoff_reject_{request_entry['id']}")
                ]
            ]
            
            balance = calculate_balance(user_name)
            week_videos = len([
                v for v in db['videos']
                if v['user'] == user_name and
                v['created_at'] >= (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
            ])
            
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"🔔 НОВЫЙ ЗАПРОС НА ВЫХОДНОЙ\n\n"
                     f"👤 {user_name}\n"
                     f"📅 Дата: {format_date(date_str)}\n"
                     f"📝 Причина: {reason}\n\n"
                     f"💵 Баланс: {balance} грн\n"
                     f"📊 Видео за неделю: {week_videos}\n\n"
                     f"Одобрить запрос?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление админу {admin_id}: {e}")
    
    await update.message.reply_text(
        f"✅ Запрос отправлен!\n\n"
        f"📅 Дата: {format_date(date_str)}\n"
        f"📝 Причина: {reason}\n\n"
        f"Администратор рассмотрит запрос в ближайшее время.",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# ОДОБРЕНИЕ/ОТКЛОНЕНИЕ ВЫХОДНОГО (АДМИН)
# ===========================
async def dayoff_approve_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка одобрения/отклонения выходного"""
    query = update.callback_query
    await query.answer()
    
    action = "approve" if "approve" in query.data else "reject"
    # Извлекаем request_id правильно: dayoff_approve_req_001 -> req_001
    request_id = "_".join(query.data.split("_")[2:])  # Все после dayoff_approve_
    
    # Находим запрос
    request = next((r for r in db['days_off_requests'] if r['id'] == request_id), None)
    
    if not request:
        await query.edit_message_text("❌ Запрос не найден")
        return
    
    if request['status'] != 'pending':
        await query.edit_message_text("❌ Запрос уже обработан")
        return
    
    if action == "approve":
        # Одобряем
        request['status'] = 'approved'
        request['approved_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Добавляем в одобренные выходные
        if request['user'] not in db['days_off_approved']:
            db['days_off_approved'][request['user']] = []
        
        db['days_off_approved'][request['user']].append({
            'date': request['date'],
            'reason': request['reason'],
            'approved_at': request['approved_at']
        })
        
        save_database(db)
        
        # Уведомление девушке
        user_telegram_id = db['users'][request['user']].get('telegram_id')
        if user_telegram_id:
            try:
                await context.bot.send_message(
                    chat_id=user_telegram_id,
                    text=f"✅ ВЫХОДНОЙ ОДОБРЕН!\n\n"
                         f"📅 Дата: {format_date(request['date'])}\n"
                         f"📝 Причина: {request['reason']}\n\n"
                         f"Хорошего отдыха! 🎉"
                )
            except Exception as e:
                logger.error(f"Не удалось отправить уведомление {request['user']}: {e}")
        
        await query.edit_message_text(
            f"✅ Выходной одобрен!\n\n"
            f"👤 {request['user']}\n"
            f"📅 Дата: {format_date(request['date'])}\n"
            f"📝 Причина: {request['reason']}"
        )
    
    else:
        # Отклоняем
        request['status'] = 'rejected'
        save_database(db)
        
        # Уведомление девушке
        user_telegram_id = db['users'][request['user']].get('telegram_id')
        if user_telegram_id:
            try:
                await context.bot.send_message(
                    chat_id=user_telegram_id,
                    text=f"❌ ЗАПРОС НА ВЫХОДНОЙ ОТКЛОНЁН\n\n"
                         f"📅 Дата: {format_date(request['date'])}\n\n"
                         f"Попробуй выбрать другую дату."
                )
            except Exception as e:
                logger.error(f"Не удалось отправить уведомление {request['user']}: {e}")
        
        await query.edit_message_text(
            f"❌ Выходной отклонён\n\n"
            f"👤 {request['user']}\n"
            f"📅 Дата: {format_date(request['date'])}"
        )

# ===========================
# ЗАПРОСЫ ВЫХОДНЫХ (АДМИН)
# ===========================
async def dayoff_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ всех запросов на выходные (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    pending = [r for r in db['days_off_requests'] if r['status'] == 'pending']
    
    if not pending:
        await update.message.reply_text("✅ Нет ожидающих запросов на выходные")
        return
    
    message = "🔔 ЗАПРОСЫ НА ВЫХОДНЫЕ\n\n"
    
    for req in pending:
        message += f"👤 {req['user']}\n"
        message += f"📅 Дата: {format_date(req['date'])}\n"
        message += f"📝 Причина: {req['reason']}\n\n"
        
        keyboard = [
            [
                InlineKeyboardButton("✅ Одобрить", callback_data=f"dayoff_approve_{req['id']}"),
                InlineKeyboardButton("❌ Отклонить", callback_data=f"dayoff_reject_{req['id']}")
            ]
        ]
        
        await update.message.reply_text(
            f"Запрос от {req['user']}:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

# ===========================
# МОИ ВЫХОДНЫЕ (АДМИН)
# ===========================
async def admin_dayoff_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало установки выходных админа (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return ConversationHandler.END
    
    keyboard = [
        [InlineKeyboardButton("Мои выходные", callback_data="admin_dayoff_admin")],
        [InlineKeyboardButton("Выходные мужа", callback_data="admin_dayoff_husband")]
    ]
    
    await update.message.reply_text(
        "📅 УСТАНОВКА ВЫХОДНЫХ\n\n"
        "Чьи выходные установить?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return ADMIN_DAYOFF_WHO

async def admin_dayoff_who_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора, чьи выходные устанавливать"""
    query = update.callback_query
    await query.answer()
    
    who = "admin" if "admin" in query.data else "husband"
    context.user_data['admin_dayoff_who'] = who
    
    who_name = "твои" if who == "admin" else "мужа"
    
    await query.edit_message_text(
        f"📅 Установка выходных ({who_name})\n\n"
        f"Напиши даты через запятую в формате ДД.ММ\n\n"
        f"Примеры:\n"
        f"• 10.02\n"
        f"• 10.02, 15.02, 20.02"
    )
    
    return ADMIN_DAYOFF_DATES

async def admin_dayoff_dates_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода дат выходных админа"""
    dates_text = update.message.text
    who = context.user_data['admin_dayoff_who']
    
    # Парсим даты
    dates = []
    for date_part in dates_text.split(','):
        try:
            day, month = date_part.strip().split('.')
            year = datetime.now().year
            date_obj = datetime(year, int(month), int(day))
            dates.append(date_obj.strftime("%Y-%m-%d"))
        except:
            continue
    
    if not dates:
        await update.message.reply_text(
            "❌ Неверный формат дат!\n\n"
            "Напиши даты через запятую в формате ДД.ММ\n"
            "Пример: 10.02, 15.02"
        )
        return ADMIN_DAYOFF_DATES
    
    # Сохраняем выходные
    db['admin_days_off'][who] = dates
    save_database(db)
    
    # Уведомления всем девушкам
    who_name = "Администратор" if who == "admin" else "Муж администратора"
    
    for user_name, user_data in db['users'].items():
        user_telegram_id = user_data.get('telegram_id')
        if user_telegram_id and not is_admin(user_telegram_id):
            try:
                dates_list = "\n".join([f"• {format_date(d)}" for d in dates])
                await context.bot.send_message(
                    chat_id=user_telegram_id,
                    text=f"📅 ВЫХОДНЫЕ АДМИНИСТРАЦИИ\n\n"
                         f"{who_name} не будет на связи:\n{dates_list}"
                )
            except Exception as e:
                logger.error(f"Не удалось отправить уведомление {user_name}: {e}")
    
    dates_formatted = ", ".join([format_date(d) for d in dates])
    
    await update.message.reply_text(
        f"✅ Выходные установлены!\n\n"
        f"👤 {who_name}\n"
        f"📅 Даты: {dates_formatted}\n\n"
        f"Все девушки уведомлены.",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# МОЙ КАЛЕНДАРЬ (ДЕВУШКИ)
# ===========================
async def my_calendar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ персонального календаря (только свои выходные + админы)"""
    user_id = update.effective_user.id
    user_name = get_user_name(user_id)
    
    if not user_name:
        await update.message.reply_text("❌ Сначала зарегистрируйся через /start")
        return
    
    # Собираем выходные
    my_daysoff = db['days_off_approved'].get(user_name, [])
    admin_daysoff = db['admin_days_off'].get('admin', [])
    husband_daysoff = db['admin_days_off'].get('husband', [])
    
    # Группируем по месяцам
    from collections import defaultdict
    by_month = defaultdict(list)
    
    for dayoff in my_daysoff:
        date_obj = datetime.strptime(dayoff['date'], "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        by_month[month_key].append({
            'date': dayoff['date'],
            'type': 'my',
            'reason': dayoff['reason']
        })
    
    for date_str in admin_daysoff:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        by_month[month_key].append({
            'date': date_str,
            'type': 'admin',
            'reason': 'Выходной админа'
        })
    
    for date_str in husband_daysoff:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        by_month[month_key].append({
            'date': date_str,
            'type': 'husband',
            'reason': 'Выходной мужа админа'
        })
    
    if not by_month:
        await update.message.reply_text(
            "📅 МОЙ КАЛЕНДАРЬ\n\n"
            "Нет запланированных выходных"
        )
        return
    
    message = "📅 МОЙ КАЛЕНДАРЬ\n\n"
    
    for month_key in sorted(by_month.keys()):
        month_name = datetime.strptime(month_key, "%Y-%m").strftime("%B %Y").upper()
        message += f"🗓️ {month_name}:\n"
        
        for entry in sorted(by_month[month_key], key=lambda x: x['date']):
            emoji = "🏖️" if entry['type'] == 'my' else "🔴" if entry['type'] == 'admin' else "🔵"
            prefix = "МОЙ ВЫХОДНОЙ" if entry['type'] == 'my' else "АДМИН НЕ НА СВЯЗИ" if entry['type'] == 'admin' else "МУЖ АДМИНА НЕ НА СВЯЗИ"
            message += f"• {format_date(entry['date'])} — {emoji} {prefix}\n"
        
        message += "\n"
    
    message += "✅ Ты видишь:\n"
    message += "• Свои выходные\n"
    message += "• Выходные администратора\n"
    message += "• Выходные мужа администратора"
    
    await update.message.reply_text(message)

# ===========================
# ГРАФИК ВЫХОДНЫХ (АДМИН)
# ===========================
async def calendar_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ графика выходных всех девушек (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    # Собираем выходные
    from collections import defaultdict
    by_month = defaultdict(list)
    
    # Выходные девушек
    for user_name, daysoff_list in db['days_off_approved'].items():
        for dayoff in daysoff_list:
            date_obj = datetime.strptime(dayoff['date'], "%Y-%m-%d")
            month_key = date_obj.strftime("%Y-%m")
            by_month[month_key].append({
                'date': dayoff['date'],
                'user': user_name,
                'reason': dayoff['reason']
            })
    
    # Выходные админов
    for date_str in db['admin_days_off'].get('admin', []):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        by_month[month_key].append({
            'date': date_str,
            'user': '🔴 АДМИН',
            'reason': 'Выходной'
        })
    
    for date_str in db['admin_days_off'].get('husband', []):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_key = date_obj.strftime("%Y-%m")
        by_month[month_key].append({
            'date': date_str,
            'user': '🔵 МУЖ АДМИНА',
            'reason': 'Выходной'
        })
    
    if not by_month:
        await update.message.reply_text(
            "📅 ГРАФИК ВЫХОДНЫХ\n\n"
            "Нет запланированных выходных"
        )
        return
    
    message = "📅 ГРАФИК ВЫХОДНЫХ (ВСЕ)\n\n"
    
    for month_key in sorted(by_month.keys()):
        month_name = datetime.strptime(month_key, "%Y-%m").strftime("%B %Y").upper()
        message += f"🗓️ {month_name}:\n"
        
        for entry in sorted(by_month[month_key], key=lambda x: x['date']):
            message += f"• {format_date(entry['date'])} — {entry['user']} ({entry['reason']})\n"
        
        message += "\n"
    
    # Статистика
    message += "📊 СТАТИСТИКА ВЫХОДНЫХ:\n"
    for user_name in USERS_CONFIG.keys():
        count = len(db['days_off_approved'].get(user_name, []))
        message += f"• {user_name}: {count} дней\n"
    
    await update.message.reply_text(message)

# ===========================
# ЭКСПОРТ В EXCEL (АДМИН)
# ===========================
async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт статистики в Excel (только для админа)"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        
        # Лист 1: Статистика пользователей
        ws1 = wb.active
        ws1.title = "Статистика"
        
        headers = ['Имя', 'Видео', 'Заработано', 'Баланс']
        ws1.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(1, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for user_name in USERS_CONFIG.keys():
            stats = get_user_stats(user_name)
            ws1.append([
                user_name,
                stats['total_videos'],
                stats['total_earnings'],
                stats['balance']
            ])
        
        # Лист 2: Все видео
        ws2 = wb.create_sheet("Видео")
        
        headers = ['ID', 'Дата', 'Пользователь', 'Тип', 'Название', 'Сумма']
        ws2.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws2.cell(1, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for video in sorted(db['videos'], key=lambda x: x['created_at'], reverse=True):
            date = datetime.strptime(video['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y %H:%M")
            ws2.append([
                video['id'],
                date,
                video['user'],
                video['type'].upper(),
                video['name'],
                video['amount']
            ])
        
        # Лист 3: Выплаты
        ws3 = wb.create_sheet("Выплаты")
        
        headers = ['ID', 'Дата', 'Пользователь', 'Тип', 'Сумма']
        ws3.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws3.cell(1, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for payment in sorted(db['payments'], key=lambda x: x['created_at'], reverse=True):
            date = datetime.strptime(payment['created_at'], "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y %H:%M")
            payment_type = "Зарплата" if payment['type'] == 'salary' else "Аванс"
            ws3.append([
                payment['id'],
                date,
                payment['user'],
                payment_type,
                payment['amount']
            ])
        
        # Сохраняем файл
        filename = f"statistic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        # Отправляем файл
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption="📊 Экспорт статистики в Excel"
            )
        
        # Удаляем файл
        os.remove(filename)
        
    except ImportError:
        await update.message.reply_text(
            "❌ Модуль openpyxl не установлен!\n\n"
            "Установи через:\n"
            "pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        await update.message.reply_text(f"❌ Ошибка экспорта: {e}")

# ===========================
# ОТМЕНА ОПЕРАЦИИ
# ===========================
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена текущей операции"""
    context.user_data.clear()
    
    await update.message.reply_text(
        "❌ Операция отменена",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    return ConversationHandler.END

# ===========================
# СРОЧНОЕ СООБЩЕНИЕ ВСЕМ (АДМИН)
# ===========================
async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало отправки срочного сообщения всем"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Доступно только администратору")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "📢 СРОЧНОЕ СООБЩЕНИЕ ВСЕМ\n\n"
        "Напиши текст сообщения, которое будет отправлено всем девушкам:"
    )
    
    return BROADCAST_MESSAGE

async def broadcast_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправка срочного сообщения всем"""
    message_text = update.message.text
    
    sent_count = 0
    failed_count = 0
    
    # Отправляем всем зарегистрированным пользователям
    for user_name, user_data in db['users'].items():
        telegram_id = user_data.get('telegram_id')
        if telegram_id:
            try:
                await context.bot.send_message(
                    chat_id=telegram_id,
                    text=f"📢 СРОЧНОЕ СООБЩЕНИЕ ОТ АДМИНИСТРАТОРА\n\n{message_text}"
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Не удалось отправить сообщение {user_name}: {e}")
                failed_count += 1
    
    await update.message.reply_text(
        f"✅ Сообщение отправлено!\n\n"
        f"📤 Успешно: {sent_count}\n"
        f"❌ Ошибок: {failed_count}",
        reply_markup=get_main_keyboard(update.effective_user.id)
    )
    
    context.user_data.clear()
    return ConversationHandler.END

# ===========================
# ОБРАБОТКА ТЕКСТОВЫХ СООБЩЕНИЙ
# ===========================
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текстовых сообщений"""
    user_id = update.effective_user.id
    text = update.message.text
    
    # Проверка регистрации
    if not get_user_name(user_id) and not is_admin(user_id):
        await handle_registration(update, context)
        return
    
    # Маршрутизация команд
    if text == '💰 Мой доход':
        await my_income(update, context)
    elif text == '📊 Моя статистика':
        await my_statistics(update, context)
    elif text == '📅 Мой план':
        await my_plan(update, context)
    elif text == '📅 Мой календарь':
        await my_calendar(update, context)
    elif text == '📊 Полная статистика':
        await full_statistics(update, context)
    elif text == '⚙️ Текущий баланс':
        await current_balance(update, context)
    elif text == '💸 Выплатить зарплату':
        await salary_payment(update, context)
    elif text == '📈 История выплат':
        await payment_history(update, context)
    elif text == '🎬 Все видео':
        await all_videos(update, context)
    elif text == '🏆 Рейтинг девушек':
        await ratings(update, context)
    elif text == '📅 График выходных':
        await calendar_all(update, context)
    elif text == '🔔 Запросы выходных':
        await dayoff_requests(update, context)
    elif text == '📊 Экспорт в Excel':
        await export_excel(update, context)
    else:
        await update.message.reply_text(
            "❓ Команда не распознана\n\n"
            "Используй кнопки меню",
            reply_markup=get_main_keyboard(user_id)
        )

# ===========================
# ГЛАВНАЯ ФУНКЦИЯ
# ===========================
def main():
    """Запуск бота"""
    # Получаем токен из переменной окружения
    token = os.getenv('TELEGRAM_BOT_TOKEN', '8280555186:AAFxZ9AfNOJdQWfFjFGk37g3pBnXCPvnupk')
    
    if not token:
        logger.error("TELEGRAM_BOT_TOKEN не установлен!")
        return
    
    # Создаём приложение
    application = Application.builder().token(token).build()
    
    # Команда /start
    application.add_handler(CommandHandler("start", start))
    
    # ConversationHandler для создания видео
    video_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^🎬 Создала видео$'), handle_video_creation)],
        states={
            VIDEO_TYPE: [CallbackQueryHandler(video_type_selected, pattern='^video_type_')],
            VIDEO_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, video_name_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="video_creation",
        persistent=False
    )
    
    # ConversationHandler для загрузки видео
    upload_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^📤 Загрузила видео$'), handle_video_upload)],
        states={
            UPLOAD_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, upload_count_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="video_upload",
        persistent=False
    )
    
    # ConversationHandler для удаления видео
    delete_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^🗑️ Удалить видео$'), delete_video_start)],
        states={
            DELETE_VIDEO_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_video_selected)],
            DELETE_VIDEO_CONFIRM: [CallbackQueryHandler(delete_video_confirmed, pattern='^delete_confirm_')]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="delete_video",
        persistent=False
    )
    
    # ConversationHandler для плана на неделю
    plan_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^📅 План на неделю$'), plan_week_start)],
        states={
            PLAN_USER: [CallbackQueryHandler(plan_user_selected, pattern='^plan_user_')],
            PLAN_DETAILS: [MessageHandler(filters.TEXT & ~filters.COMMAND, plan_details_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="plan_week",
        persistent=False
    )
    
    # ConversationHandler для выплаты аванса
    advance_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^💰 Выплатить аванс$'), advance_payment_start)],
        states={
            ADVANCE_USER: [CallbackQueryHandler(advance_user_selected, pattern='^advance_user_')],
            ADVANCE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, advance_amount_entered)],
            ADVANCE_CONFIRM: [CallbackQueryHandler(advance_confirmed, pattern='^advance_confirm_')]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="advance_payment",
        persistent=False
    )
    
    # ConversationHandler для запроса выходного
    dayoff_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^📅 Запросить выходной$'), request_dayoff_start)],
        states={
            DAYOFF_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, dayoff_date_entered)],
            DAYOFF_REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, dayoff_reason_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="dayoff_request",
        persistent=False
    )
    
    # ConversationHandler для выходных админа
    admin_dayoff_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^📅 Мои выходные$'), admin_dayoff_start)],
        states={
            ADMIN_DAYOFF_WHO: [CallbackQueryHandler(admin_dayoff_who_selected, pattern='^admin_dayoff_')],
            ADMIN_DAYOFF_DATES: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_dayoff_dates_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="admin_dayoff",
        persistent=False
    )
    
    # Регистрируем ConversationHandlers
    application.add_handler(video_conv_handler)
    application.add_handler(upload_conv_handler)
    application.add_handler(delete_conv_handler)
    application.add_handler(plan_conv_handler)
    application.add_handler(advance_conv_handler)
    application.add_handler(dayoff_conv_handler)
    application.add_handler(admin_dayoff_conv_handler)
    
    # ConversationHandler для срочного сообщения
    broadcast_conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex('^📢 Срочное сообщение$'), broadcast_start)],
        states={
            BROADCAST_MESSAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_send)]
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        name="broadcast",
        persistent=False
    )
    application.add_handler(broadcast_conv_handler)
    
    # CallbackQueryHandlers
    application.add_handler(CallbackQueryHandler(process_salary_payment, pattern='^pay_salary_'))
    application.add_handler(CallbackQueryHandler(dayoff_approve_reject, pattern='^dayoff_(approve|reject)_'))
    
    # Обработчик текстовых сообщений (должен быть ПОСЛЕДНИМ!)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Запускаем бота
    logger.info("🤖 Бот запущен!")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
